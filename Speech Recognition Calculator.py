import os
import re 
import tempfile
import openpyxl

try:
  import speech_recognition as sr
except:
  !pip install SpeechRecognition
  import speech_recognition as sr
try:
  import pyaudio
except:
  !pip install pyaudio
  import pyaudio

try:
  from gtts import gTTS
except:
  !pip install gTTS
  from gtts import gTTS
    
try:
  import pygame 
except:
  !pip install pygame  
  import pygame 
    
from io import BytesIO 

class speech_to_text:
  def __init__(self):  
    self.rg = sr.Recognizer()
  def listen(self,lang='zh-tw'):  
    with sr.Microphone() as source:
      audioData = self.rg.listen(source)
      try:
        text = self.rg.recognize_google(audioData, language=lang)    
      except:
        text = ''
    return text

class text_to_speech:
  def __init__(self):
    self.active_mp3  = 'c:\\tmp\\tmp2.mp3'
    pygame.mixer.init()
  def __del__(self):
    try:
      os.unlink(self.active_mp3)  
    except:
      pass  
  def speak(self,text,lang='zh-tw'): 
    tts= gTTS(text, lang=lang)
    tts.save(self.active_mp3)
    pygame.mixer.music.load(self.active_mp3)
    pygame.mixer.music.play()
    while pygame.mixer.music.get_busy():
      continue
    pygame.mixer.music.unload()
    return
  

def find_price(produce):
    # 要處理的Excel檔案名稱
      wb = openpyxl.load_workbook('menu.xlsx')
      sheet = wb.worksheets[0]
      price_updates_dict = {produce}
      findproduce = 0

      #使用for loop掃描所有A欄品項，如果比對一致，就回傳單價
      for rowNum in range(2, sheet.max_row, 1):
          produceName = sheet.cell(rowNum, 1).value
          if produceName in price_updates_dict:
              findproduce = 1
              return sheet.cell(rowNum, 2).value

st = speech_to_text()
ts = text_to_speech()    
total = 0
s = st.listen()

#每個品項的份數
d={"甜不辣":0,"薯條":0,"鹹酥雞":0,"四季豆":0,"百頁豆腐":0,"杏鮑菇":0} 
#前面目前只找到判斷一個字的方法，品名第一個字和最後一個字可能都要不一樣
sfind = re.compile(r'(["甜不辣""薯條""鹹酥雞""四季豆""百頁豆腐""杏鮑菇"]\d+[份])|(\d+[份]["甜不辣""薯條""鹹酥雞""四季豆""百頁豆腐""杏鮑菇"])')  
for m in sfind.finditer(s):  
    a = m.group()  
    print(a)
    name=re.findall(r'["甜不辣""薯條""鹹酥雞""四季豆""百頁豆腐""杏鮑菇"]',a)  
    n = re.findall(r'\d+',a)  
    if name[0]=="甜" or name[0]=="辣":
        name[0]="甜不辣"
    elif name[0]=="薯" or name[0]=="條":
        name[0]="薯條"
    elif name[0]=="鹹" or name[0]=="雞":
        name[0]="鹹酥雞"
    elif name[0]=="四" or name[0]=="豆":
        name[0]="四季豆"
    elif name[0]=="百" or name[0]=="腐":
        name[0]="百頁豆腐"
    elif name[0]=="杏" or name[0]=="菇":
        name[0]="杏鮑菇"
    d[name[0]]+= int(n[0])  


for i,j in d.items(): 
    if j != 0:
        total += int(find_price(i)) * j

ts.speak(str(total))